import requests
import time
import random
import re
import threading
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import os
from selenium.webdriver.common.keys import Keys
import numpy as np
from bezier.curve import Curve
import math
from selenium.webdriver.common.action_chains import ActionChains
# Import class di chuyển chuột giống người
from human_mouse import HumanMouse

import subprocess
import ctypes
import sys
#Chạy file .bat với quyền administrator
def run_as_admin(bat_file):
    """
    Chạy file .bat với quyền administrator
    """
    try:
        print(f"Đang chạy file {bat_file} với quyền administrator...")
        if ctypes.windll.shell32.IsUserAnAdmin():
            # Đã có quyền admin rồi, chạy trực tiếp
            subprocess.run(bat_file, shell=True, check=True)
        else:
            # Cần yêu cầu quyền admin và chạy lại lệnh
            ctypes.windll.shell32.ShellExecuteW(None, "runas", "cmd.exe", f"/c {bat_file}", None, 1)
        print(f"File {bat_file} đã được chạy với quyền administrator.")
        return True
    except Exception as e:
        print(f"Lỗi khi chạy file {bat_file}: {e}")
        return False
# ---------------------------------------
# Thiết lập đường dẫn động
# ---------------------------------------
# Lấy đường dẫn thư mục hiện tại (nơi script đang chạy)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Lấy đường dẫn thư mục cha (nơi chứa profiles.xlsx và proxy.txt)
PARENT_DIR = os.path.dirname(SCRIPT_DIR)

# Đường dẫn đến các file trong thư mục cha
PROFILES_FILE = os.path.join(PARENT_DIR, 'profiles.xlsx')
PROXY_FILE = os.path.join(PARENT_DIR, 'proxy.txt')

# Đường dẫn đến các file trong thư mục script
HASHTAG_FILE = os.path.join(SCRIPT_DIR, 'hashtag.txt')
MUSIC_FILE = os.path.join(SCRIPT_DIR, 'clidtiktok.txt')
USE_VIDEO_FILE = os.path.join(SCRIPT_DIR, 'use_video.txt')
USE_MUSIC_FILE = os.path.join(SCRIPT_DIR, 'use_music.txt')
PROFILE_VIDEO_FILE = os.path.join(SCRIPT_DIR, 'profile_video.txt')
SUCCESS_UPLOAD_FILE = os.path.join(SCRIPT_DIR, 'successupload.txt')
PROFILE_ERROR_FILE = os.path.join(SCRIPT_DIR, 'profileloi.txt')

# ---------------------------------------
# Khởi tạo các lock
# ---------------------------------------
profile_index_lock = threading.Lock()
excel_lock = threading.Lock()
music_lock = threading.Lock()
successupload_lock = threading.Lock()
profile_video_lock = threading.Lock()
use_video_lock = threading.Lock()
use_music_lock = threading.Lock()
reset_lock = threading.Lock() # Lock cho việc reset danh sách

# ---------------------------------------
# Hàm cập nhật proxy cho profile
# ---------------------------------------

def update_proxy(profile_id, raw_proxy):
    """
    Cập nhật proxy cho profile trước khi mở.
    Gửi POST request với raw_proxy (prefix là "http://").
    Nếu trả về "Profile not found" thì log profile id ra file profileloi.txt.
    Trả về {"success": True} nếu thành công, ngược lại trả về {"success": False}.
    """
    update_url = f"http://127.0.0.1:19995/api/v3/profiles/update/{profile_id}"
    headers = {"accept": "application/json", "Content-Type": "application/json"}
    data = {"raw_proxy": f"{raw_proxy}"}
    try:
        r = requests.post(update_url, headers=headers, json=data)
        r.raise_for_status()  # Raise HTTPError for bad responses (4xx or 5xx)

        response_json = r.json()
        if response_json.get("success") and response_json.get("message") == "Update profile success":
            print(f"Proxy updated successfully for profile {profile_id}.")
            return {"success": True}
        elif (not response_json.get("success")) and response_json.get("message") == "Profile not found":
            print(f"Update failed. Profile not found: {profile_id}")
            with open("profileloi.txt", "a") as f:
                f.write(str(profile_id) + "\n")
            return {"success": False}
        else:
            print(f"Unexpected response when updating proxy for profile {profile_id}: {response_json}")
            return {"success": False}
    except requests.exceptions.RequestException as e:  # Catch network errors
        print(f"Exception updating proxy for profile {profile_id}: {e}")
        return {"success": False}
    except ValueError as e:  # Catch JSON decoding errors
        print(f"Exception decoding JSON response for profile {profile_id}: {e}")
        print(f"Response text: {r.text}")  # Print the raw response
        return {"success": False}
    except Exception as e:  # Catch all other exceptions
        print(f"Unexpected error updating proxy for profile {profile_id}: {e}")
        return {"success": False}

# ---------------------------------------
# Hàm đổi IP
# ---------------------------------------
def change_ip(change_ip_url):
    """
    Gọi GET request tới API đổi IP.
    Nếu API trả về lỗi "Vui lòng chờ sau X giây" thì đợi (X+2) giây rồi retry.
    """
    while True:
        try:
            r = requests.get(change_ip_url)
            try:
                response_json = r.json()
            except Exception:
                print("Lỗi khi parse JSON từ API đổi IP, retry sau 10 giây...")
                time.sleep(10)
                continue

            if response_json.get("status") == "success":
                print("Đổi IP thành công.")
                return True
            elif response_json.get("status") == "error":
                error_msg = response_json.get("error", "")
                m = re.search(r"Vui lòng chờ sau (\d+) giây", error_msg)
                if m:
                    wait_seconds = int(m.group(1)) + 2
                    print(f"Lỗi đổi IP: {error_msg}. Đợi {wait_seconds} giây rồi retry lại...")
                    time.sleep(wait_seconds)
                    continue
                else:
                    print(f"Phản hồi lỗi không xác định từ API đổi IP: {response_json}")
                    return False
            else:
                print(f"Phản hồi không mong đợi từ API đổi IP: {response_json}")
                return False
        except Exception as e:
            print(f"Exception while changing IP: {e}. Đợi 10 giây rồi retry lại...")
            time.sleep(10)
            continue

# ---------------------------------------
# Đọc file proxy.txt
# ---------------------------------------
try:
    with open(PROXY_FILE, "r") as f:
        proxy_lines = [line.strip() for line in f.readlines()]
    if len(proxy_lines) != 6:
        raise ValueError("Expected 4 proxy lines in proxy.txt.")

    proxies = []
    for line in proxy_lines:
        proxy_parts = line.split("|")
        if len(proxy_parts) != 2:
            raise ValueError(f"Invalid format in proxy.txt.\nExpected: proxy|changeIP_url. Line: {line}")
        raw_proxy = proxy_parts[0]
        change_ip_url = proxy_parts[1]
        proxies.append({"raw_proxy": raw_proxy, "change_ip_url": change_ip_url})
        print(f"Loaded proxy: {raw_proxy}, Change IP URL: {change_ip_url}")

except Exception as e:
    print(f"Error reading proxy.txt: {e}")
    proxies = []

# ---------------------------------------
# Đọc file Excel profiles.xlsx
#  - Cột A: Profile ID
#  - Cột F: Kết quả (có thể chứa "Thành Công")
# Chỉ lấy profile nếu cột F là "Thành Công"
# ---------------------------------------
workbook = load_workbook(PROFILES_FILE)
worksheet = workbook.active

red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

profiles = []
try:
    with successupload_lock:
        with open(SUCCESS_UPLOAD_FILE, "r") as f:
            success_profile_ids = [line.strip() for line in f.readlines()]
except FileNotFoundError:
    success_profile_ids = []
    
for i, row in enumerate(worksheet.iter_rows(min_row=2, max_col=6, values_only=False), start=2):
    profile_id = row[0].value  # Cột A
    result_value = worksheet.cell(row=i, column=6).value  # Cột F

    # Bỏ qua nếu profile đã upload thành công
    if str(profile_id) in success_profile_ids:
        print(f"Skipping profile {profile_id} (row {i}) vì đã upload thành công rồi.")
        continue

    # Chỉ xử lý nếu cột F là "Thành Công"
    if result_value and str(result_value).strip() == "Thành Công":
        profiles.append({
            "id": profile_id,
            "row": i
        })
    else:
        print(f"Skipping profile {profile_id} (row {i}) vì cột F không phải 'Thành Công'.")

# Biến toàn cục để xử lý đa luồng
profile_index = 0  # index profile hiện tại

# ---------------------------------------
# Các hàm hỗ trợ
# ---------------------------------------
def load_list(filename):
    try:
        with open(filename, "r", encoding="utf-8") as f:
            return [line.strip() for line in f.readlines()]
    except FileNotFoundError:
        return []

def save_list(filename, data):
    try:
        with open(filename, "w", encoding="utf-8") as f:
            for item in data:
                f.write(item + "\n")
    except Exception as e:
        print(f"Lỗi khi ghi vào file {filename}: {e}")

def get_random_item(data, used_file, profile_id, item_type="video"):
    """
    Chọn ngẫu nhiên một item từ danh sách data, tránh trùng lặp.
    """

    for attempt in range(3):  # Thử 3 lần
        used_items = load_list(used_file) # Đọc lại danh sách đã dùng ở mỗi lần thử
        profile_used_videos = []

        if not data: # Check if the original data list is empty
            print(f"Danh sách {item_type} đầu vào trống. Reset danh sách: {used_file}")
            with reset_lock: # Chỉ một luồng được phép reset
                if item_type == "video":
                    lock_to_use = use_video_lock
                else:
                    lock_to_use = use_music_lock

                with lock_to_use:
                    save_list(used_file, [])  # Reset use_video.txt or use_music.txt
            return None  # No items available after reset either.

        if item_type == "video":
            try:
                with profile_video_lock: # Đảm bảo an toàn khi đọc profile_video.txt
                    with open(PROFILE_VIDEO_FILE, "r", encoding="utf-8") as f:
                        for line in f:
                            line = line.strip()
                            if not line:  # Bỏ qua dòng trống
                                continue
                            if ":" not in line:  # Kiểm tra xem dòng có dấu ":" hay không
                                print(f"Warning: Skipping invalid line in profile_video.txt: {line}")
                                continue  # Bỏ qua dòng này nếu không hợp lệ
                            try:
                                p_id, video = line.split(":", 1)
                            except ValueError as e:
                                print(f"Error splitting line: {line} - {e}")
                                continue  # Bỏ qua dòng này nếu có lỗi

                            if p_id == str(profile_id):
                                profile_used_videos.append(video)
            except FileNotFoundError:
                pass  # Không có file profile_video.txt, không sao cả

            available_items = [
                item
                for item in data
                if item not in used_items and item not in profile_used_videos
            ]
        elif item_type == "music":
            available_items = [item for item in data if item not in used_items]
        else:
            print(f"Unknown item_type: {item_type}")
            return None

        if not available_items:
            print(f"Đã sử dụng hết item khả dụng. Reset danh sách: {used_file}")
            with reset_lock:  # Chỉ một luồng được phép reset
                if item_type == "video":
                    lock_to_use = use_video_lock
                else:
                    lock_to_use = use_music_lock

                with lock_to_use:
                    save_list(used_file, [])  # Reset use_video.txt or use_music.txt
            continue  # Thử lại sau khi reset
            
        if available_items:
            item = random.choice(available_items)
            used_items.append(item)

            if item_type == "video":
                lock_to_use = use_video_lock
            else:
                lock_to_use = use_music_lock

            with lock_to_use:
                save_list(used_file, used_items)

            if item_type == "video":
                # Lưu thông tin video đã dùng cho profile
                try:
                    with profile_video_lock: # Đảm bảo an toàn khi ghi vào profile_video.txt
                        with open(PROFILE_VIDEO_FILE, "a", encoding="utf-8") as f:
                            f.write(f"{profile_id}:{item}\n")
                except Exception as e:
                    print(f"Error writing to profile_video.txt: {e}")

            return item
        else:
            print(f"Không có item khả dụng sau lần thử {attempt + 1}.")

    print("Không tìm thấy item khả dụng sau nhiều lần thử.")
    return None

def type_like_human(element, text):
    for char in text:
        element.send_keys(char)
        time.sleep(random.uniform(0.05, 0.1))

# ---------------------------------------
# Thread function
# ---------------------------------------
def process_profile(thread_id, proxy_data, window_pos):
    global profile_index

    raw_proxy = proxy_data["raw_proxy"]
    change_ip_url = proxy_data["change_ip_url"]

    while True:
        # Lấy index profile an toàn
        with profile_index_lock:
            if profile_index >= len(profiles):
                print(f"Thread {thread_id}: No more profiles to process.")
                break
            current_profile_index = profile_index
            profile_index += 1

        profile = profiles[current_profile_index]
        profile_id = profile["id"]
        row_number = profile["row"]

        # Log số profile còn lại trước khi xử lý
        remaining_profiles = len(profiles) - current_profile_index - 1
        print(f"Thread {thread_id}: Processing profile {profile_id} (Row {row_number}).  {remaining_profiles} profiles remaining.")


        print(f"Thread {thread_id}: Processing profile {profile_id} (Row {row_number})")

        # 1) Cập nhật proxy
        if raw_proxy and change_ip_url:
            if not update_proxy(profile_id, raw_proxy):
                print(f"Thread {thread_id}: Skipping profile {profile_id} due to proxy update failure.")
                with excel_lock:
                    worksheet.cell(row=row_number, column=6).value = "Lỗi Cập Nhật Proxy"
                    for cell in worksheet[row_number]:
                        cell.fill = red_fill
                    workbook.save(PROFILES_FILE)
                continue

            # 2) Đổi IP
            if not change_ip(change_ip_url):
                print(f"Thread {thread_id}: Sự cố đổi IP, không loại bỏ profile {profile_id}, sẽ retry.")
        else:
            print(f"Thread {thread_id}: Proxy info not available; skipping proxy update.")

        # 3) Mở profile qua API
        start_url = f"http://127.0.0.1:19995/api/v3/profiles/start/{profile_id}?addination_args=--lang%3Dvi&win_pos={window_pos}&win_size=1800%2C1080&win_scale=0.35"
        print(f"Thread {thread_id}: Opening profile via URL: {start_url}")
        try:
            start_resp = requests.get(start_url)
            start_resp.raise_for_status()
        except Exception as e:
            print(f"Thread {thread_id}: Error opening profile {profile_id}: {e}")
            continue

        start_data = start_resp.json()
        if not start_data.get("success"):
            print(f"Thread {thread_id}: Failed to open profile {profile_id}: {start_data}")
            continue

        driver_path = start_data.get("data", {}).get("driver_path")
        remote_debugging_address = start_data.get("data", {}).get("remote_debugging_address")
        browser_location = start_data.get("data", {}).get("browser_location")

        if not driver_path or not remote_debugging_address:
            print(f"Thread {thread_id}: Missing driver_path or remote_debugging_address, skipping profile.")
            continue

        # 4) Khởi tạo Selenium
        options = Options()
        # options.binary_location = browser_location # nếu cần
        options.add_experimental_option("debuggerAddress", remote_debugging_address)
        service = Service(executable_path=driver_path)
        try:
            driver = webdriver.Chrome(service=service, options=options)
            human_mouse = HumanMouse(driver, window_size=(1800, 1080), window_scale=0.35)
        except Exception as e:
            print(f"Thread {thread_id}: Error initializing webdriver for profile {profile_id}: {e}")
            continue

        # ---------------------------
        # 5) Truy cập trang Upload
        # ---------------------------
        wait = WebDriverWait(driver, 30)
        tiktok_url = "https://www.tiktok.com/tiktokstudio/upload"
        print(f"Thread {thread_id}: Navigating to {tiktok_url}")
        try:
            driver.get(tiktok_url)
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
            print(f"Thread {thread_id}: TikTok upload page loaded.")
        except Exception as e:
            print(f"Thread {thread_id}: Error loading TikTok upload page for profile {profile_id}: {e}")
            close_profile(driver, profile_id)
            continue

        # Kiểm tra captcha
        try:
            captcha_element = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div#captcha-verify-container-main-page"))
            )
            print(f"Thread {thread_id}: Captcha xuất hiện, đợi biến mất...")
            WebDriverWait(driver, 30).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div#captcha-verify-container-main-page"))
            )
            print(f"Thread {thread_id}: Captcha đã biến mất.")
        except:
            print(f"Thread {thread_id}: Không thấy captcha hoặc không biến mất sau 30s, tiếp tục...")

        # Kiểm tra đăng xuất
        try:
            login_modal = driver.find_element(By.CSS_SELECTOR, "h2#login-modal-title")
            print(f"Thread {thread_id}: Tài khoản bị đăng xuất.")
            close_and_update_excel(driver, profile_id, row_number, "Tài Khoản Bị Đăng Xuất", "Tài khoản TikTok đã bị đăng xuất.")
            continue
        except:
            print(f"Thread {thread_id}: Tài khoản vẫn đăng nhập.")

        # ---------------------------
        # 6) Upload video
        # ---------------------------
        video_dir = "D:\\tiktok"
        videos = [f for f in os.listdir(video_dir) if os.path.isfile(os.path.join(video_dir, f))]
        if not videos:
            print(f"Thread {thread_id}: Không có video nào trong thư mục D:\\tiktok")
            close_profile(driver, profile_id)
            continue

        video_path = get_random_item(videos, USE_VIDEO_FILE, profile_id, item_type="video")
        if not video_path:
            print(f"Thread {thread_id}: Không tìm thấy video hợp lệ.")
            close_profile(driver, profile_id)
            continue

        full_video_path = os.path.join(video_dir, video_path)
        print(f"Thread {thread_id}: Chọn video: {full_video_path}")

        try:
            upload_input = driver.find_element(By.XPATH, "//input[@type='file']")
            upload_input.send_keys(full_video_path)
            print(f"Thread {thread_id}: Video {video_path} đã được chọn.")
        except Exception as e:
            print(f"Thread {thread_id}: Lỗi upload video {video_path}: {e}")
            close_profile(driver, profile_id)
            continue

        # ---------------------------
        # 7) Nhập hashtag
        # ---------------------------
        time.sleep(2)
        hashtags = load_list(HASHTAG_FILE)
        if not hashtags:
            print(f"Thread {thread_id}: Không có hashtag trong file hashtag.txt")
            close_profile(driver, profile_id)
            continue

        try:
            caption_editor = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.public-DraftStyleDefault-block")))
            for _ in range(3):
                hashtag = random.choice(hashtags)
                type_like_human(caption_editor, hashtag)
                time.sleep(2)
                caption_editor.send_keys(Keys.ENTER)
        except Exception as e:
            print(f"Thread {thread_id}: Lỗi nhập hashtag: {e}")
            close_profile(driver, profile_id)
            continue

        # ---------------------------
        # 8) Chờ video upload xong
        # ---------------------------
        try:
            # Tạo một WebDriverWait mới với thời gian chờ là 60 giây
            upload_wait = WebDriverWait(driver, 60)
            upload_wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "[data-testid='CheckCircleFill'] > svg")))
            print(f"Thread {thread_id}: Video đã upload xong.")
        except:
            print(f"Thread {thread_id}: Video không upload xong sau 60 giây.")
            close_profile(driver, profile_id)
            continue

        # ---------------------------
        # 9) Chọn nhạc và đăng video
        # ---------------------------
        music_name = None  # Initialize music_name
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            post_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.TUXButton")))
            human_mouse.human_move_to_element(post_button, click=True)
            time.sleep(random.uniform(0.6, 1.2))
            print(f"Thread {thread_id}: Đã bấm nút đăng video.")
        except Exception as e:
            print(f"Thread {thread_id}: Lỗi bấm nút đăng video: {e}")
            close_profile(driver, profile_id)
            continue

        try:
            search_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input.search-bar-input")))
            with music_lock:  # Thêm lock ở đây
                music_list = load_list(MUSIC_FILE)
            if not music_list:
                print(f"Thread {thread_id}: Không có nhạc trong file clidtiktok.txt")
                close_profile(driver, profile_id)
                continue

            music_name = get_random_item(music_list, USE_MUSIC_FILE, profile_id, item_type="music")

            if not music_name:
                print(f"Thread {thread_id}: Không tìm thấy nhạc hợp lệ.")
                close_profile(driver, profile_id)
                continue

            human_mouse.human_move_to_element(search_input, click=True)
            type_like_human(search_input, music_name)
            search_input.send_keys(Keys.ENTER)
            time.sleep(random.uniform(0.6, 1.2))

            first_music_card = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.jsx-915372485.music-card-container:first-child")))
            human_mouse.human_move_to_element(first_music_card, click=False)
            first_music_card.click()
            time.sleep(random.uniform(0.6, 1.2))
            music_operation = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.jsx-915372485.music-card-operation")))
            human_mouse.human_move_to_element(music_operation, click=False)
            music_operation.click()

            time.sleep(random.uniform(0.8, 1.2))

            confirm_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.TUXButton--primary")))
            human_mouse.human_move_to_element(confirm_button, click=True)

            time.sleep(random.uniform(0.6, 1.2))

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)

            final_post_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".jsx-1810272162 > .Button__root--type-primary")))
            human_mouse.human_move_to_element(final_post_button, click=True)

        except Exception as e:
            print(f"Thread {thread_id}: Lỗi chọn nhạc và đăng video: {e}")
            close_profile(driver, profile_id)
            continue

        # ---------------------------
        # 10) Xử lý modal spam (nếu có)
        # ---------------------------
        try:
            modal = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.jsx-1150920910.common-modal-body")))
            print(f"Thread {thread_id}: Modal spam xuất hiện.")
            appeal_buttons = driver.find_elements(By.CSS_SELECTOR, "button.TUXButton.TUXButton--default.TUXButton--medium.TUXButton--primary.appeal-button")
            if len(appeal_buttons) >= 2:
                human_mouse.human_move_to_element(appeal_buttons[0], click=True)
                time.sleep(random.uniform(0.8, 1.3))
                human_mouse.human_move_to_element(appeal_buttons[1], click=True)
                print(f"Thread {thread_id}: Đã bấm nút kháng nghị.")

            with excel_lock:
                worksheet.cell(row=row_number, column=6).value = "Tài Khoản Bị Spam"
                for cell in worksheet[row_number]:
                    cell.fill = red_fill
                workbook.save(PROFILES_FILE)
        except:
            print(f"Thread {thread_id}: Không có modal spam.")

        # Lưu profile ID đã upload thành công và music name
        try:
            with successupload_lock:
                with open(SUCCESS_UPLOAD_FILE, "a") as f:
                    f.write(str(profile_id) + "\n")
        except Exception as e:
            print(f"Error writing to successupload.txt: {e}")

        print(f"Thread {thread_id}: Đã đăng video thành công. Lưu profile ID vào successupload.txt")

        close_profile(driver, profile_id)
        #break # kết thúc vòng lặp sau khi hoàn thành 1 profile

def close_and_update_excel(driver, profile_id, row_number, result, message):
    print(f"Closing profile {profile_id} and updating Excel: {message}")
    with excel_lock:
        worksheet.cell(row=row_number, column=6).value = result
        for cell in worksheet[row_number]:
            cell.fill = red_fill
        workbook.save(PROFILES_FILE)
    close_profile(driver, profile_id)

def close_profile(driver, profile_id):
    close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
    print(f"Closing profile with URL: {close_url}")
    try:
        requests.get(close_url)
        print(f"Profile {profile_id} closed successfully.")
    except Exception as e:
        print(f"Error closing profile {profile_id}: {e}")


# ---------------------------------------
# Main
# ---------------------------------------
if __name__ == "__main__":
    # Kiểm tra đủ 6 proxy chưa
    if len(proxies) < 6:
        print("Không đủ 6 proxy để chạy 6 luồng, vui lòng kiểm tra lại proxy.txt")
        exit(1)

    # Load data
    hashtags = load_list(HASHTAG_FILE)
    music_list = load_list(MUSIC_FILE)

    # Tạo 6 luồng, mỗi luồng xài 1 proxy
    thread1 = threading.Thread(target=process_profile, args=(1, proxies[0], "0,0"))
    thread2 = threading.Thread(target=process_profile, args=(2, proxies[1], "1800,0"))
    thread3 = threading.Thread(target=process_profile, args=(3, proxies[2], "3600,0"))
    thread4 = threading.Thread(target=process_profile, args=(4, proxies[3], "0,1080"))
    thread5 = threading.Thread(target=process_profile, args=(5, proxies[4], "1800,1080"))
    thread6 = threading.Thread(target=process_profile, args=(6, proxies[5], "3600,1080"))

    thread1.start()
    thread2.start()
    thread3.start()
    thread4.start()
    thread5.start()
    thread6.start()

    thread1.join()
    thread2.join()
    thread3.join()
    thread4.join()
    thread5.join()
    thread6.join()

    print("\nCompleted processing all profiles.")
    # Chạy file login.bat sau khi hoàn tất
    login_bat_path = r"C:\Users\namhuunamsv\Desktop\tiktok\login\login.bat"
    run_as_admin(login_bat_path)
